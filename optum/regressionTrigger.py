# Databricks notebook source
# DBTITLE 1,Installation
# MAGIC %pip install psycopg2
# MAGIC %pip install adlfs
# MAGIC %pip install azure-identity
# MAGIC %pip install azure-keyvault-secrets
# MAGIC %pip install openpyxl
# MAGIC %pip install azure-identity azure-storage-blob
# MAGIC %pip install azure-storage-blob azure-mgmt-datafactory
# MAGIC %pip install azure-monitor-query
# MAGIC %pip install azure-identity azure-mgmt-monitor
# MAGIC
# MAGIC

# COMMAND ----------

dbutils.widgets.text("environment", "qa")
env = dbutils.widgets.get("environment")
print(f"environment : {env}")

# COMMAND ----------

# DBTITLE 1,Function Calls
import os
import logging
import psycopg2
import pandas.io.sql as sqlio
from azure.identity import DefaultAzureCredential
from azure.keyvault.secrets import SecretClient
from pyspark.sql import SparkSession


dap_nonprod_sql_pwd_nm = 'dap-nonprod-postgresql-pe-password'
dap_nonprod_emailer_URL_nm = 'dap-nonprod-logic-app-http-url'
key_vault_name = 'dap-nonprod-kv'

DB_CONFIG = {
    "user": "eip_dap_admin",
    "host": "dap-nonprod-postgresql-pe.postgres.database.azure.com",
    "port": 5432,
    "database": f"{env}_dap_run_log"
}

# -----------------------------------
# Helpers
# -----------------------------------

def get_secret(secret_name, key_vault_name):
    credential = DefaultAzureCredential()
    client = SecretClient(vault_url=f"https://{key_vault_name}.vault.azure.net/", credential=credential)
    return client.get_secret(secret_name).value

def get_db_connection():
    password = get_secret(dap_nonprod_sql_pwd_nm, key_vault_name)
    return psycopg2.connect(
        user=DB_CONFIG["user"],
        password=password,
        host=DB_CONFIG["host"],
        port=DB_CONFIG["port"],
        database=DB_CONFIG["database"]
    )

def execute_query(query):
    try:
        with get_db_connection() as conn:
            df = sqlio.read_sql_query(query, conn)
            return df
    except Exception as e:
        logging.error(f"Query execution failed: {e}")
        raise

def read_sql_file(file_path):
    if not os.path.exists(file_path):
        logging.error(f"SQL file not found: {file_path}")
        return None
    with open(file_path, 'r') as file:
        return file.read()

# -----------------------------------
# Core Functionalities
# -----------------------------------

def run_post_test_sql(sql_file_path, test_id):
    query = read_sql_file(sql_file_path)
    if not query:

        return "FAIL"
    try:
        print(query)
        df = execute_query(query)
        if df.success_flag[0] == 1:
            return "PASS"
        else:
            return "FAIL"
    except Exception as e:
        logging.error(f"Error in post-test SQL for {test_id}: {e}")
        return "FAIL"

def get_rerun_count(query):
    try:
        return execute_query(query)
    except Exception as e:
        logging.error(f"Failed to get rerun count: {e}")
        return None

def get_error_result_from_sql(sql_file_path):
    query = read_sql_file(sql_file_path)
    error_message = ""
    if not query:
        return None
    try:
        df = execute_query(query)
        if df.empty:
            error = "SQL result is empty"
        elif 'error_message' not in df.columns:
            error = "'error_message' column missing in SQL result"
        else:
            raw_error = df.iloc[0]['error_message']
            if raw_error and len(raw_error) > 0:
                error = raw_error[0]
            else:
                error = "Error message is empty or invalid"
        return error
    except Exception as e:
        logging.error(f"Failed to fetch error result: {e}")
        return None

# -----------------------------------
# Parquet/CSV Read Utilities
# -----------------------------------

def read_parquet_count_rows(file_path):
    df = spark.read.parquet(file_path)
    return df.count()

def read_parquet_file_count(file_path):
    df = spark.read.parquet(file_path)
    return len(df.inputFiles())

def read_source_row_count(file_path):
    if file_path.endswith('.csv') or file_path.endswith('.txt') or file_path.endswith('.dat'):
        df = spark.read.csv(file_path, header=True)
    return df.count()
# -----------------------------------
# Output Management
# -----------------------------------

outputs = {}

def save_output(test_id, description, result, error="null"):
    outputs[test_id] = {
        "description": description,
        "result": result,
        "error": error
    }


# COMMAND ----------

# DBTITLE 1,Read Json
import json
with open(f'/dbfs/mnt/data/{env}/Test_files/regression/test_config.json', 'r') as f:
    data = f.read()
jsonObject = json.loads(data)
test_case_path = jsonObject.get('test_case_path').format(env=env)
test_cases = jsonObject.get('test_cases')
dbfs_testcase_path = "/dbfs"+test_case_path
test_cases_dict = {}
for test_case in test_cases:  
    test_cases_dict = {test_case.get('TestID'): test_case for test_case in test_cases}
display(test_cases_dict)

# COMMAND ----------

# DBTITLE 1,Generic Test case
import pandas as pd
from openpyxl import load_workbook
# A test case execution function for generic testcases
def run_test_case(test_case, test_case_path):
    test_id = test_case.get('TestID')
    id = test_case.get("ID")
    post_sql = test_case.get('PostTest_SQL')
    input_file = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    description = test_case.get('Test_Case_Description')
    error_sql = test_case.get("ErrorSQL")

    sql_path = f'{dbfs_testcase_path}/{post_sql}'
    print(sql_path)
    result = run_post_test_sql(sql_path, test_id)
    try:
        if result == "PASS":

            input_path = f'{test_case_path}/tc{id}/input_files/{input_file}'
            if 'xlsx' in input_file:
                input_path = f'{dbfs_testcase_path}/tc{id}/input_files/{input_file}'
                wb = load_workbook(filename=input_path,data_only=True)
                sheet = wb.active
                source_count = sum(1 for row in sheet.iter_rows() if any(cell.value is not None for cell in row)) - 1 #as the excel file having header row
            else:
                source_count = read_source_row_count(input_path)
            parquet_file_count = read_parquet_file_count(landed_folder)
            parquet_row_count = read_parquet_count_rows(landed_folder)
            if parquet_file_count == 1 and source_count == parquet_row_count:
                display(f"Testcase {test_id} Passed")
                save_output(test_id, description, "PASS")
            else:
                error = get_error_result_from_sql(f'{test_case_path}/{error_sql}')
                display(f"Testcase {test_id} Failed: {error}")
                save_output(test_id, description, "FAIL", error)

        else:
            logging.error(f"Testcase {test_id} Failed due to post-test SQL")
            save_output(test_id, description, "FAIL", "Post-test SQL failed")
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        error = str(e)
        save_output(test_id, description, "FAIL", error)



# COMMAND ----------

# DBTITLE 1,TC-PUSH-001
test_case = test_cases_dict.get("TC_PUSH_001")
run_test_case(test_case, test_case_path)

# COMMAND ----------

# DBTITLE 1,TC-PUSH-002
test_case = test_cases_dict.get("TC_PUSH_002")
run_test_case(test_case, test_case_path)

# COMMAND ----------

# DBTITLE 1,TC-PUSH-003
test_case = test_cases_dict.get("TC_PUSH_003")
run_test_case(test_case, test_case_path)

# COMMAND ----------

# DBTITLE 1,TC-PUSH-004
test_case = test_cases_dict.get("TC_PUSH_004")
run_test_case(test_case, test_case_path)

# COMMAND ----------

# DBTITLE 1,TC-PUSH-005
test_case = test_cases_dict.get("TC_PUSH_005")
run_test_case(test_case, test_case_path)

# COMMAND ----------

# DBTITLE 1,TC-PUSH-005a
test_case = test_cases_dict.get("TC_PUSH_005a")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder')
    test_case_desc = test_case.get('Test_Case_Description')
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    error_sql = test_case.get("ErrorSQL")
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    # sql_result = run_sql_file(sql_file_path, post_test_sql, test_id)
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_005a:", sql_result)
    try:
        target_counts = {}
        for i in range(1, 3):
            target_path = landed_folder.format(env= env , part_count=i)
            target_counts[f'part{i}_count'] = read_parquet_count_rows(target_path)
        print(target_counts)
        if sql_result == 'PASS' and len(target_counts.keys()) == 2:
            display("Testcase TC_PUSH_005a Passed")
            save_output("TC_PUSH_005a",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_005a Failed")
            save_output("TC_PUSH_005a", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e)) 

# COMMAND ----------

# DBTITLE 1,TC-PUSH-006
test_case = test_cases_dict.get("TC_PUSH_006")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder')
    test_case_desc = test_case.get('Test_Case_Description')
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    error_sql = test_case.get("ErrorSQL")
    source_names = test_case.get('Sources')
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_006:", sql_result)
    try:
        if sql_result == "PASS":
            source_names = ['claims' , 'payers' , 'devices']
            source_count = {}
            for source in source_names: 
                file_path = f'{test_case_path}/tc6/source_files/*{source}*'
                source_count[source] = spark.read.csv(file_path,header=True).count()
            display(source_count)

            target_count = {}
            for source in source_names: 
                target_path = landed_folder.format(env= env ,source=source)
                target_count[source] = read_parquet_count_rows(target_path)
            display(target_count)

            src_tgt_row_cnt_validation = True 

            if source_count.keys() == target_count.keys():
                for key in source_count:
                    if source_count[key] == target_count[key]:
                        pass
                    else:
                        src_tgt_row_cnt_validation = False
                        break
                    
        if sql_result == "PASS" and src_tgt_row_cnt_validation:
            display("Testcase TC_PUSH_006 Passed")
            save_output("TC_PUSH_006",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_006 Failed")
            save_output("TC_PUSH_006", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC-PUSH-009
test_case = test_cases_dict.get("TC_PUSH_009")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    error_sql = test_case.get("ErrorSQL")
    test_case_desc = test_case.get('Test_Case_Description')  

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_009:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc9/input_files/{test_file_names}'

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)  

        if sql_result == "PASS" and parquet_file_count == 1 and parquet_row_count == 0:
            display("Testcase TC_PUSH_009 Passed")
            save_output("TC_PUSH_009",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_009 Failed")
            save_output("TC_PUSH_009", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC-PUSH-010
test_case = test_cases_dict.get("TC_PUSH_010")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    error_sql = test_case.get("ErrorSQL")
    test_case_desc = test_case.get('Test_Case_Description')  

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_010:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc10/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)  

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_010 Passed")
            save_output("TC_PUSH_010",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_010 Failed")
            save_output("TC_PUSH_010",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_011
test_case = test_cases_dict.get("TC_PUSH_011")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    error_sql = test_case.get("ErrorSQL")
    test_case_desc = test_case.get('Test_Case_Description')  

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_011:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc11/input_files/{test_file_names}' 

        if sql_result == "PASS":
            display("Testcase TC_PUSH_011 Passed")
            save_output("TC_PUSH_011",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_011 Failed")
            save_output("TC_PUSH_011", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_012
test_case = test_cases_dict.get("TC_PUSH_012")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    error_sql = test_case.get("ErrorSQL")
    test_case_desc = test_case.get('Test_Case_Description')  

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_012:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc12/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)  

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_012 Passed")
            save_output("TC_PUSH_012", test_case_desc,"PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_012 Failed")
            save_output("TC_PUSH_012",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_013
test_case = test_cases_dict.get("TC_PUSH_013")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')
    error_sql = test_case.get("ErrorSQL")  

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)    
    print("Query execution status of TC_PUSH_013:", sql_result)
    try:
        if sql_result == "PASS":
            display("Testcase TC_PUSH_013 Passed")
            save_output("TC_PUSH_013",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_013 Failed")
            save_output("TC_PUSH_013",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_014
test_case = test_cases_dict.get("TC_PUSH_014")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description') 
    error_sql = test_case.get("ErrorSQL") 

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_014:", sql_result)
    try:

        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc14/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count) 

            # centura_epsi_monthly_sp / tarnishedbronze / processed_files / bad_row_data  verification is pending 

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_014 Passed")
            save_output("TC_PUSH_014",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_014 Failed")
            save_output("TC_PUSH_014",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_015
test_case = test_cases_dict.get("TC_PUSH_015")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description') 
    error_sql = test_case.get("ErrorSQL") 

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_015:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc15/input_files/{test_file_names}'

            # comparing source manifest csv and generated minifest csv in tanshied folder is pending 

        if sql_result == "PASS":
            display("Testcase TC_PUSH_015 Passed")
            save_output("TC_PUSH_015",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_015 Failed")
            save_output("TC_PUSH_015",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))


# COMMAND ----------

# DBTITLE 1,TC_PUSH_016
test_case = test_cases_dict.get("TC_PUSH_016")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description') 
    error_sql = test_case.get("ErrorSQL") 

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_016:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc16/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count) 

            # Check the email for bad rows. Check the bad row file generated for the 3 errors. pending

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_016 Passed")
            save_output("TC_PUSH_016",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_016 Failed")
            save_output("TC_PUSH_016", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_017
test_case = test_cases_dict.get("TC_PUSH_017")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description') 
    error_sql = test_case.get("ErrorSQL") 

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_017:", sql_result)
    try:

        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc17/input_files/{test_file_names}'

        if sql_result == "PASS":
            display("Testcase TC_PUSH_017 Passed")
            save_output("TC_PUSH_017",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_017 Failed")
            save_output("TC_PUSH_017",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_018
test_case = test_cases_dict.get("TC_PUSH_018")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description') 
    error_sql = test_case.get("ErrorSQL") 

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_018:", sql_result)
    try:
        if sql_result == "PASS":
            display("Testcase TC_PUSH_018 Passed")
            save_output("TC_PUSH_018",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_018 Failed")
            save_output("TC_PUSH_018", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_019
test_case = test_cases_dict.get("TC_PUSH_019")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description') 
    error_sql = test_case.get("ErrorSQL") 

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_019:", sql_result)
    try:
        if sql_result == "PASS":
            display("Testcase TC_PUSH_019 Passed")
            save_output("TC_PUSH_019",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_019 Failed")
            save_output("TC_PUSH_019",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))


# COMMAND ----------

# DBTITLE 1,TC_PUSH_026
test_case = test_cases_dict.get("TC_PUSH_026")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_026:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc26/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count) 

        if sql_result == "PASS" and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_026 Passed")
            save_output("TC_PUSH_026", test_case_desc,"PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_026 Failed")
            save_output("TC_PUSH_026", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_027
test_case = test_cases_dict.get("TC_PUSH_027")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_027:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc27/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count) 

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_027 Passed")
            save_output("TC_PUSH_027",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_027 Failed")
            save_output("TC_PUSH_027",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_029
test_case = test_cases_dict.get("TC_PUSH_029")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_029:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc29/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count) 

            listOfFiles = dbutils.fs.ls(landed_folder)
            for file in listOfFiles:
                fileSize = file.size
                print(f"File: {file.name}, Size: {fileSize} bytes")
                if fileSize < 512000000:
                    sizeValidation = "PASS"
                else:
                    sizeValidation = "FAIL"      
            print("Parquet file count in folder:", len(listOfFiles))
            print("sizeValidation:", sizeValidation)

        if sql_result == "PASS" and source_row_count == parquet_row_count and sizeValidation == "PASS":
            display("Testcase TC_PUSH_029 Passed")
            save_output("TC_PUSH_029",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_029 Failed")
            save_output("TC_PUSH_029",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_030
test_case = test_cases_dict.get("TC_PUSH_030")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_030:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc30/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_030 Passed")
            save_output("TC_PUSH_030", test_case_desc,"PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_030 Failed")
            save_output("TC_PUSH_030", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_034
from azure.identity import DefaultAzureCredential
from azure.mgmt.datafactory import DataFactoryManagementClient
from azure.mgmt.datafactory.models import RunFilterParameters
from datetime import datetime, timedelta

test_case = test_cases_dict.get("TC_PUSH_034")
print(test_case)
test_id = test_case.get('TestID')
test_file_names = test_case.get('Test_File_Names')
test_file_names = 'TC034_devices_20240619_part_0001.dat'
test_case_desc = test_case.get('Test_Case_Description')
feed_name = test_case.get('Feed_Name')

subscription_id = "80085be5-acec-4711-a455-717ff9d02c08"
resource_group_name = "eip-hsodev-rg"
factory_name = "eip-hsodev-datafactory"

credential = DefaultAzureCredential()
adf_client = DataFactoryManagementClient(credential, subscription_id)
end_time = datetime.now()
start_time = end_time - timedelta(minutes=270)

filter_params = RunFilterParameters(
    last_updated_after=start_time,
    last_updated_before=end_time
)
pipeline_runs = adf_client.pipeline_runs.query_by_factory(
    resource_group_name,
    factory_name,
    filter_params
)
try:
    count = 0
    for run in pipeline_runs.value:
        if run.pipeline_name == f"hsodev-{env}-InvokingDVP-etl-pipeline" and run.parameters['FeedName'] == feed_name  :
            print(run)
            if run.status == 'Succeeded':
                count += 1
    if count >= 15 :
        display("Testcase TC_PUSH_034 Passed")
        save_output("TC_PUSH_034", test_case_desc,"PASS")
    else:
        display("Testcase TC_PUSH_034 Failed")
        save_output("TC_PUSH_034", test_case_desc,"FAIL","DVP trigger failed ")
except Exception as e:
    print(f"Testcase {test_id} Failed due to exception: {e}")
    save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_038
test_case = test_cases_dict.get("TC_PUSH_038")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    error_sql = test_case.get("ErrorSQL")
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_038:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc38/input_files/{test_file_names}'
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)
        if sql_result == "PASS" and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_038 Passed")
            save_output("TC_PUSH_038",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_038 Failed")
            save_output("TC_PUSH_038", test_case_desc,"FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_39
test_case = test_cases_dict.get("TC_PUSH_039")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_039:", sql_result)
    try:
        if sql_result == "PASS":
            display("Testcase TC_PUSH_039 Passed")
            save_output("TC_PUSH_039",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_039 Failed")
            save_output("TC_PUSH_039",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_040
test_case = test_cases_dict.get("TC_PUSH_040")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_040:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc40/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_040 Passed")
            save_output("TC_PUSH_040", test_case_desc,"PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_040 Failed")
            save_output("TC_PUSH_040",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_041
test_case = test_cases_dict.get("TC_PUSH_041")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_041:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc41/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_041 Passed")
            save_output("TC_PUSH_041",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_041 Failed")
            save_output("TC_PUSH_041",test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))
        

# COMMAND ----------

# DBTITLE 1,TC_PUSH_042
test_case = test_cases_dict.get("TC_PUSH_042")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description') 
    error_sql = test_case.get("ErrorSQL") 

    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_042:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc42/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_042 Passed")
            save_output("TC_PUSH_042", test_case_desc,"PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_042 Failed")
            save_output("TC_PUSH_042", test_case_desc,"FAIL",error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_043
test_case = test_cases_dict.get("TC_PUSH_043")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    print(test_case_desc)
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    rerun_count_sql = test_case.get('PreTest_SQL')
    Archive_path = test_case.get('Archive_Folder')
    error_sql = test_case.get("ErrorSQL")

    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_043:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc43/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)

            # Get rerun count from database
            re_run_count_df = get_rerun_count(rerun_count_sql)
            re_run_count = re_run_count_df.iloc[0]['re_run_count']
            print("re_run_count_df: ", re_run_count)

            # Read archive folder using rerun count and count parquet files in archive folder 
            archive_file_count = read_parquet_file_count(Archive_path.format(re_run_count = re_run_count,env=env))
            print("archive_file_count: ", archive_file_count)

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count and archive_file_count == 1:
            display("Testcase TC_PUSH_043 Passed")
            save_output("TC_PUSH_043",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_043 Failed")
            save_output("TC_PUSH_043",test_case_desc, "FAIL",error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_044
test_case = test_cases_dict.get("TC_PUSH_044")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder').format(env=env)
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path,test_id)
    print("Query execution status of TC_PUSH_044:", sql_result)
    try:
        if sql_result == "PASS":
            file_path = f'{test_case_path}/tc44/input_files/{test_file_names}'

            # Read source file with multiple extension support and count rows
            source_row_count = read_source_row_count(file_path)
            print("source_row_count: ", source_row_count)

            # Read folder and count parquet files
            parquet_file_count = read_parquet_file_count(landed_folder)
            print("parquet_file_count: ", parquet_file_count)

            # Read parquet file and count rows
            parquet_row_count = read_parquet_count_rows(landed_folder)
            print("parquet_row_count: ", parquet_row_count)

            # Archive folder check, Extracted_holding folder check,Auto Feed completion status mail check--Pending

        if sql_result == "PASS" and parquet_file_count == 1 and source_row_count == parquet_row_count:
            display("Testcase TC_PUSH_044 Passed")
            save_output("TC_PUSH_044",test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_044 Failed")
            save_output("TC_PUSH_044",test_case_desc, "FAIL",error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_045
from azure.identity import DefaultAzureCredential
from azure.mgmt.datafactory import DataFactoryManagementClient
from azure.mgmt.datafactory.models import RunFilterParameters
from datetime import datetime, timedelta

test_case = test_cases_dict.get("TC_PUSH_045")
post_test_sql = test_case.get('PostTest_SQL')
test_id = test_case.get('TestID')
test_file_names = test_case.get('Test_File_Names')
landed_folder = test_case.get('Landed_Folder').format(env=env)
test_case_desc = test_case.get('Test_Case_Description') 
error_sql = test_case.get("ErrorSQL")
subscription_id = "80085be5-acec-4711-a455-717ff9d02c08"
resource_group_name = "eip-hsodev-rg"
factory_name = "eip-hsodev-datafactory"

credential = DefaultAzureCredential()
adf_client = DataFactoryManagementClient(credential, subscription_id)
end_time = datetime.now()
start_time = end_time - timedelta(minutes=90)

filter_params = RunFilterParameters(
    last_updated_after=start_time,
    last_updated_before=end_time
)
pipeline_runs = adf_client.pipeline_runs.query_by_factory(
    resource_group_name,
    factory_name,
    filter_params
)
try:
    cnt = 0
    for run in pipeline_runs.value:
        if run.pipeline_name == f"hsodev-{env}-DAP_ADF_Orchestration-etl-pipeline" and run.parameters['FeedName']  == 'centura_daily_autotest':
            cnt+=1
    if cnt > 20:
        display("Testcase TC_PUSH_045 Passed")
        save_output("TC_PUSH_045", test_case_desc,"PASS")
    else:
        display("Testcase TC_PUSH_045 Failed")
        save_output("TC_PUSH_045", test_case_desc,"FAIL",f"{cnt} pipelines triggered in last 60 minutes")
except Exception as e:
    print(f"Testcase {test_id} Failed due to exception: {e}")
    save_output(test_id, test_case_desc, "FAIL", str(e))
    

# COMMAND ----------

# DBTITLE 1,TC_PUSH_046
test_case = test_cases_dict.get("TC_PUSH_046")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    test_case_desc = test_case.get('Test_Case_Description')  
    error_sql = test_case.get("ErrorSQL")
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    print(test_case_desc)
    print(sql_file_path)
    sql_result = run_post_test_sql(sql_file_path,test_id)
    validator = True
    sources = test_case.get("Sources")
    print("Query execution status of TC_PUSH_046:", sql_result)
    try:
        if sql_result == "PASS":
            targetPaths = test_case.get('Landed_Folder')
            sourcePaths = test_case.get('Source_Path')
            results = {}
            for source in sources:
                results[source] = {}
                parts_count = 3 if source == "claims" else 2
                for i in range(1, parts_count + 1):
                    results[source][f"part{i}"] = {}
                    tgt_directory = targetPaths[source].format(env=env) + str(i)
                    results[source][f"part{i}"]["target_count"] = read_parquet_count_rows(tgt_directory)
                    src_directory = sourcePaths[source].format(env=env , i=i)
                    results[source][f"part{i}"]["source_count"] = read_source_row_count(src_directory)
            print(results)
            for source, parts in results.items():
                for key, value in parts.items():
                    if value["target_count"] != value["source_count"]:
                        validator = False
        if sql_result == "PASS" and validator:
            display("Testcase TC_PUSH_046 Passed")
            save_output("TC_PUSH_046", test_case_desc, "PASS")
        else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_046 Failed")
            save_output("TC_PUSH_046", test_case_desc, "FAIL", error)
    except Exception as e:
        print(f"Testcase {test_id} Failed due to exception: {e}")
        save_output(test_id, test_case_desc, "FAIL", str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_047
test_case = test_cases_dict.get("TC_PUSH_047")
post_test_sql = test_case.get('PostTest_SQL')
test_id = test_case.get('TestID')
test_file_names = test_case.get('Test_File_Names')
landed_folder = test_case.get('Landed_Folder').format(env=env)
test_case_desc = test_case.get('Test_Case_Description')
error_sql = test_case.get("ErrorSQL")

sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
print(test_case_desc)
# Call run_sql_file() function to get test result
sql_result = run_post_test_sql(sql_file_path,test_id)
try:
    if sql_result == "PASS":
        display("Testcase TC_PUSH_047 Passed")
        save_output("TC_PUSH_047", test_case_desc,"PASS")
    else:
            sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
            error=get_error_result_from_sql(sql_error_file_path)
            display("Testcase TC_PUSH_047 Failed")
            save_output("TC_PUSH_047",test_case_desc, "FAIL",error)
except Exception as e:
    print(f"Testcase {test_id} Failed due to exception: {e}")
    save_output(test_id, test_case_desc, "FAIL", str(e))
    

# COMMAND ----------

# DBTITLE 1,TC_PUSH_048

test_case = test_cases_dict.get("TC_PUSH_048")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    test_case_desc = test_case.get('Test_Case_Description')
    error_sql = test_case.get("ErrorSQL")
sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
print(test_case_desc)
# Call run_sql_file() function to get test result
sql_result = run_post_test_sql(sql_file_path,test_id)
parts = [2,3]
targetPath = test_case.get('Landed_Folder')
try : 
    for part in parts : 
        tgt_directory = targetPath.format(part=part,env=env)
        read_parquet_count_rows(tgt_directory)
except:
    display("Testcase TC_PUSH_048 Failed")
    save_output("TC_PUSH_048",test_case_desc, "FAIL","the expecting part files are not there in the landed path") 
if sql_result == "PASS":
    display("Testcase TC_PUSH_048 Passed")
    save_output("TC_PUSH_048", test_case_desc,"PASS")
else:
        sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
        error=get_error_result_from_sql(sql_error_file_path)
        display("Testcase TC_PUSH_048 Failed")
        save_output("TC_PUSH_048",test_case_desc, "FAIL",error)   
    



# COMMAND ----------

# DBTITLE 1,TC_PUSH_049
from azure.identity import DefaultAzureCredential
from azure.mgmt.datafactory import DataFactoryManagementClient
from azure.mgmt.datafactory.models import RunFilterParameters
from datetime import datetime, timedelta
import json

test_case = test_cases_dict.get("TC_PUSH_049")
if test_case:
    test_id = test_case.get('TestID')
    file_names_to_check = test_case.get('Test_File_Names')
    test_case_desc = test_case.get('Test_Case_Description')  
    subscription_id = "80085be5-acec-4711-a455-717ff9d02c08"
    resource_group_name = "eip-hsodev-rg"
    factory_name = "eip-hsodev-datafactory"
    credential = DefaultAzureCredential()
    adf_client = DataFactoryManagementClient(credential, subscription_id)
    end_time = datetime.now()
    start_time = end_time - timedelta(minutes=1200)
    filter_params = RunFilterParameters(
    last_updated_after=start_time,
    last_updated_before=end_time
    )
    pipeline_runs = adf_client.pipeline_runs.query_by_factory(
    resource_group_name,
    factory_name,
    filter_params
    )
    missing_files = []
    try:
        for run in pipeline_runs.value:
            if run.pipeline_name == f"hsodev-{env}-TriggerFAV-etl-pipeline" and run.status == "Succeeded":
                parameters = run.parameters
                fav_filelist = json.loads(parameters['FAVFilelist'])
                file_names_in_run = [file['file_name'] for file in fav_filelist]
                print([file['file_name'] for file in fav_filelist])
                missing_files = [file for file in file_names_to_check if file not in file_names_in_run]
        if not missing_files:
            display("All files are present.")
            save_output("TC_PUSH_049",test_case_desc, "PASS")                
        else:
            display(f"Missing files: {missing_files}")
            save_output("TC_PUSH_049",test_case_desc, "FAIL",f"fav trigger failed for the files : {missing_files}")
    except Exception as e:
        display(f"Error: {e}")
        save_output("TC_PUSH_049",test_case_desc, "FAIL",str(e))

# COMMAND ----------

# DBTITLE 1,TC_PUSH_050
import logging
from azure.monitor.query import LogsQueryClient, LogsQueryStatus
from azure.identity import DefaultAzureCredential
from datetime import timedelta
import re

# Configuration
CONFIG = {
    "app_insights_app_id": "9031317c-052e-4f0d-a7a6-a8c73b3049ba",
    "timespan_hours": 2
}
error = None

def get_ticket_numbers(app_insights_app_id, timespan):
    try:
        credential = DefaultAzureCredential()
        client = LogsQueryClient(credential)

        query = """
        AppTraces
        | where Message contains "ticket"
        | project TimeGenerated, Message
        """

        response = client.query_workspace(
            workspace_id=app_insights_app_id,
            query=query,
            timespan=timespan
        )

        if response.status == LogsQueryStatus.SUCCESS:
            ticket_numbers = []
            for table in response.tables:
                for row in table.rows:
                    message = row[1]  # Access the 'Message' field
                    match = re.search(r"\bUS\d+\b", message)
                    if match:
                        ticket_numbers.append(match.group(0))
            return ticket_numbers
        else:
            error = f"Application Insights query failed: {response.status}"
            return None
    except Exception as e:
        error = f"An unexpected error occurred: {e}"
        return None


if __name__ == "__main__":
    test_case = test_cases_dict.get("TC_PUSH_050")
    test_id = test_case.get('TestID')
    file_names_to_check = test_case.get('Test_File_Names')
    test_case_desc = test_case.get('Test_Case_Description') 
    timespan = timedelta(hours=CONFIG["timespan_hours"])
    ticket_numbers = get_ticket_numbers(CONFIG["app_insights_app_id"], timespan)
    if ticket_numbers:
        print("Ticket Numbers Found:")
        for num in ticket_numbers:
            display(num)
        save_output("TC_PUSH_050",test_case_desc, "PASS")

    else:
        error = "No ticket numbers found or an error occurred."
        display('no rally ticket created')
        save_output("TC_PUSH_050",test_case_desc, "FAIL",error)


# COMMAND ----------

# DBTITLE 1,TC_PUSH_053
test_case = test_cases_dict.get("TC_PUSH_053")
if test_case:
    post_test_sql = test_case.get('PostTest_SQL')
    test_id = test_case.get('TestID')
    test_file_names = test_case.get('Test_File_Names')
    landed_folder = test_case.get('Landed_Folder')
    source_folder = test_case.get('Source_Folder')
    test_case_desc = test_case.get('Test_Case_Description')  
    sql_file_path = f'{dbfs_testcase_path}/{post_test_sql}'
    error_sql = test_case.get("ErrorSQL")
    print(test_case_desc)
    file_path = test_case.get("sorce_folder")

    # Call run_sql_file() function to get test result
    sql_result = run_post_test_sql(sql_file_path, test_id)

    print("Query execution status of TC_PUSH_053:", sql_result)
    if sql_result == "PASS":    
        try:
            source_counts = {}
            i = 1
            for file_name in test_file_names:
                source_counts[i] = read_source_row_count(file_path=file_path.format(dbfs_testcase_path = dbfs_testcase_path , file_name = file_name))
                i += 1
            target_counts = {}
            for part in range(1,3):
                target_counts[part] = read_parquet_count_rows(file_path=landed_folder.format(env = env ,part=part)) 

        except Exception as e:
            print(f"An error occurred: {e}")
        if sql_result == "PASS" and source_counts == target_counts :
            display("Testcase TC_PUSH_053 Passed")
            save_output("TC_PUSH_053", test_case_desc,"PASS")
    else:
        sql_error_file_path = f'{dbfs_testcase_path}/{error_sql}'
        res=get_error_result_from_sql(sql_error_file_path)
        display("Testcase TC_PUSH_053 Failed")
        save_output("TC_PUSH_053", test_case_desc,"FAIL", error)

# COMMAND ----------

# DBTITLE 1,Testcase Report
import csv  
from datetime import datetime  
  
# Define the current timestamp  
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')  
  
# Define the path to save the CSV report with timestamp suffix  
csv_report_path = f'/dbfs/mnt/data/{env}/Test_files/regression/test_case_report_{timestamp}.csv'  

with open(csv_report_path, mode='w', newline='') as file:  
    writer = csv.writer(file)  
    writer.writerow(['Test Case',  'Status', 'Description' , 'Error Message'])  
    for test_case, status in outputs.items():  
        writer.writerow([test_case, status['result'], status['description'], status['error']])  
  
# Display the CSV file path  
df = spark.read.csv(f'dbfs:/mnt/data/{env}/Test_files/regression/test_case_report_{timestamp}.csv', header=True) 
ordered_df = df.orderBy("Test Case") 
display(ordered_df)

# COMMAND ----------

total_tests = df.count()
passed_tests = df.filter(df["Status"] == 'PASS').count()
failed_tests = df.filter(df["Status"] == 'FAIL').count()
pass_percentage = (passed_tests / total_tests) * 100

# COMMAND ----------

print(pass_percentage)

# COMMAND ----------

import pandas as pd
from datetime import datetime, timezone, timedelta
import requests



def create_email_body_from_csv_pandas(ordered_df):
    """Creates an email body with a CSV file as a formatted table using pandas."""
    try:
        df = ordered_df.toPandas()
        html_table = df.to_html(index=False, justify='center', border=1) #index=False removes row index,justify center aligns cells
        email_body = f"""<h2>Please find the below test case report :</h2>\n <p>Total Test Cases: {total_tests}</p>\n<p>Passed Test Cases: {passed_tests}</p>\n<p>Failed Test Cases: {failed_tests}</p>\n<p>Pass Percentage: {pass_percentage:.2f}%</p>\n\n{html_table}""" # Add HTML header
        return email_body
    except FileNotFoundError:
        print(f"Error: CSV file not found at {ordered_df}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

body = create_email_body_from_csv_pandas(ordered_df)

dap_nonprod_emailer_URL = get_secret(dap_nonprod_emailer_URL_nm, key_vault_name)
today = datetime.now().astimezone().astimezone(timezone(timedelta(hours=5, minutes=30))).strftime("%Y-%m-%d %H:%M")
subject = f"DAP regression automation {env.upper()} Test Case Report - {today}"
recipients = "alavala_babu@optum.com;ss@optum.com"#jsonObject.get('Recipients')
sender = "odedap@omail.o360.cloud"
json_data = {"message": body,"subject": subject, "receiver":recipients, "from": sender} 

output = requests.post(dap_nonprod_emailer_URL, json=json_data)
print(output.status_code)







# Import necessary libraries
import psycopg2
import pandas as pd
from datetime import datetime

# --------- CONFIGURATION ---------
client_name = 'regressinv1'
feed_name = 'centuar_monthly_autitest'
test_case_id = "TC_PUSH_ONCE_PER_CADENCE"
test_case_desc = "Validate once-per-cadence email notification for centuar_monthly_autitest"
current_date = datetime.today().date()

# --------- DB CONNECTION ---------
def get_postgres_connection():
    return psycopg2.connect(
        host="your_host",             # ⬅️ Replace with actual host
        database="your_database",     # ⬅️ Replace with actual DB
        user="your_user",             # ⬅️ Replace with actual username
        password="your_password",     # ⬅️ Replace with actual password
        port="5432"                   # ⬅️ Change if using non-default port
    )

# --------- FETCH FILE ARRIVAL DATA ---------
def fetch_file_arrival_status(client_name, feed_name):
    query = """
        SELECT * FROM file_arrival_status 
        WHERE client_name = %s 
          AND feed_name = %s 
          AND notify_type = 'once_per_cadence'
    """
    conn = get_postgres_connection()
    df = pd.read_sql(query, conn, params=(client_name, feed_name))
    conn.close()
    return df

# --------- VALIDATE EMAIL CONTENT ---------
def validate_email(response_text, template):
    if not response_text:
        return False
    for keyword in template.get("keywords", []):
        if keyword.lower() not in response_text.lower():
            return False
    return True

# --------- DUMMY EMAIL FETCH (REPLACE IN PROD) ---------
def get_logicapp_email_response(client_name, feed_name):
    # Replace with actual Logic App API / Log Analytics logic
    data = {
        "response": [
            "Feed completion status: incomplete. is_missing: true. Missed_files: 3. Expectation_date: 2025-07-28. Feed_details: centuar_monthly_autitest. Arrival_date: 2025-07-28"
        ]
    }
    return pd.DataFrame(data)

# --------- RESULT LOGGER ---------
def save_output(test_case_id, description, status, reason=None):
    log = {
        "TestCaseID": test_case_id,
        "Description": description,
        "Status": status,
        "Reason": reason or ""
    }
    display(log)

# --------- MAIN TEST CASE EXECUTION ---------
try:
    file_arrival_df = fetch_file_arrival_status(client_name, feed_name)

    # Convert date columns
    file_arrival_df['yellow_email_start_date'] = pd.to_datetime(file_arrival_df['yellow_email_start_date'], errors='coerce')
    file_arrival_df['red_email_start_date'] = pd.to_datetime(file_arrival_df['red_email_start_date'], errors='coerce')

    # Filter for today's date match
    eligible_df = file_arrival_df[
        (file_arrival_df['yellow_email_start_date'].dt.date == current_date) |
        (file_arrival_df['red_email_start_date'].dt.date == current_date)
    ]

    if not eligible_df.empty:
        print("✅ Eligible record(s) found. Proceeding to email validation...")

        # Define expected email template
        email_template = {
            "keywords": [
                "feed completion status", "incomplete", "is_missing",
                "missed_files", "expectation_date", "feed_details", "arrival_date"
            ]
        }

        logicapp_response_df = get_logicapp_email_response(client_name, feed_name)
        response_text = logicapp_response_df["response"].iloc[0] if not logicapp_response_df.empty else ""

        if validate_email(response_text, email_template):
            save_output(test_case_id, test_case_desc, "PASS")
        else:
            save_output(test_case_id, test_case_desc, "FAIL", "Missing keyword(s) in email content")
    else:
        print("⏭️ No matching cadence for today's date. Test skipped.")
        save_output(test_case_id, test_case_desc, "SKIPPED", "No matching cadence today")

except Exception as e:
    save_output(test_case_id, test_case_desc, "ERROR", str(e))



def run_daily_email_check_test():
    client_name = "regressinv1"
    test_case_id = "TC_PUSH_YELLOW_RED_EMAIL"
    test_case_desc = "Validate yellow and red emails for missing files"
    current_date = datetime.today().date()

    df = fetch_active_email_windows(client_name)

    if df.empty:
        save_output(test_case_id, test_case_desc, "SKIPPED", "No yellow/red window active today")
        return

    # Template for both email types
    yellow_template = {
        "keywords": ["incomplete", "missing", "expected", "yellow", "feed_name", "arrival_date"]
    }
    red_template = {
        "keywords": ["escalation", "alert", "red", "critical", "feed_name", "arrival_date"]
    }

    yellow_pass, red_pass = True, True  # Default to True, fail only if triggered and not matched

    # Check for yellow email window
    yellow_df = df[
        (df['yellow_email_start_date'].dt.date <= current_date) &
        (df['yellow_email_end_date'].dt.date >= current_date)
    ]
    if not yellow_df.empty:
        yellow_email_df = get_logicapp_email_response(client_name, "YELLOW")
        yellow_response = yellow_email_df["response"].iloc[0] if not yellow_email_df.empty else ""
        yellow_pass = validate_email(yellow_response, yellow_template)

    # Check for red email window
    red_df = df[
        (df['red_email_start_date'].dt.date <= current_date) &
        (df['red_email_end_date'].dt.date >= current_date)
    ]
    if not red_df.empty:
        red_email_df = get_logicapp_email_response(client_name, "RED")
        red_response = red_email_df["response"].iloc[0] if not red_email_df.empty else ""
        red_pass = validate_email(red_response, red_template)

    # Final decision
    if yellow_pass and red_pass:
        save_output(test_case_id, test_case_desc, "PASS")
    else:
        if not yellow_pass and not red_pass:
            reason = "Both yellow and red email validation failed"
        elif not yellow_pass:
            reason = "Yellow email validation failed"
        else:
            reason = "Red email validation failed"
        save_output(test_case_id, test_case_desc, "FAIL", reason)
